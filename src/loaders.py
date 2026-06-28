"""
Platform-specific data loaders for author business analytics.
Each function reads raw export files and returns normalized DataFrames.
"""

import calendar
import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from src.schemas import ACX_NEW_FORMAT, DROP_COLUMNS_COMMON


def extract_date_from_acx_filename(filename):
    """
    Parse month and year from ACX report filename.
    Handles both formats:
    - New: Royalty_..._MONTHLY_Apr_2026.xlsx (abbreviated)
    - Old: S_D_Simper_ACX_MONTHLY_April_2021.xls (full month name)
    
    Returns: datetime object for the 1st of that month
    """
    months = {}
    for i, name in enumerate(calendar.month_name[1:], 1):
        months[name.lower()] = i
    for i, abbr in enumerate(calendar.month_abbr[1:], 1):
        months[abbr.lower()] = i

    pattern = r'_([A-Za-z]{3,9})_(\d{4})\.xlsx?$'
    match = re.search(pattern, filename, re.IGNORECASE)

    if match:
        month_str, year = match.groups()
        month_num = months.get(month_str.lower())
        if month_num:
            return datetime(int(year), month_num, 1)

    raise ValueError(f"Cannot parse date from filename: {filename}")


def _match_titles_to_catalog(df, catalog, threshold=0.8):
    """
    Match catalog entries based on book title similarity (fallback when IDs unavailable).
    
    Args:
        df: Sales DataFrame
        catalog: BookCatalog instance
        threshold: Min similarity ratio (0-1) for fuzzy matching
    
    Returns:
        DataFrame enriched with catalog info where matches found
    """
    if catalog.raw_catalog is None:
        df['series'] = None
        df['canonical_work_slug'] = None
        df['edition_format'] = None
        return df

    from difflib import SequenceMatcher

    def find_best_match(title, candidates):
        """Find closest match from candidate titles."""
        if pd.isna(title) or not isinstance(title, str):
            return None, 0

        title_lower = title.lower().strip()
        title_clean = re.sub(r'\s*:\s*.*$', '', title_lower)
        title_clean = re.sub(r'\s*\([^)]*\)', '', title_clean)
        title_clean = title_clean.replace('&', 'and').strip()

        best_score = 0
        best_match = None

        for idx, candidate in enumerate(candidates):
            if pd.isna(candidate) or not isinstance(candidate, str):
                continue

            cand_lower = candidate.lower().strip()
            cand_clean = re.sub(r'\s*:\s*.*$', '', cand_lower)
            cand_clean = re.sub(r'\s*\([^)]*\)', '', cand_clean)
            cand_clean = cand_lower.replace('&', 'and').strip()

            score = SequenceMatcher(None, title_clean, cand_clean).ratio()

            if score > best_score:
                best_score = score
                best_match = idx

        if best_score >= threshold:
            return best_match, best_score
        return None, best_score

    catalog_titles = catalog.raw_catalog['display_title'].tolist()
    catalog_info = catalog.raw_catalog[['series', 'canonical_work_slug', 'edition_format']]

    df = df.copy()
    df['_match_result'] = df['book_title'].apply(lambda t: find_best_match(t, catalog_titles))

    def apply_single_match(row):
        idx, score = row['_match_result']
        if idx is not None:
            return pd.Series([
                catalog_info.loc[idx, 'series'],
                catalog_info.loc[idx, 'canonical_work_slug'],
                catalog_info.loc[idx, 'edition_format']
            ])
        return pd.Series([None, None, None])

    matched_cols = df.apply(apply_single_match, axis=1)
    df[['series', 'canonical_work_slug', 'edition_format']] = matched_cols

    df = df.drop(columns=['_match_result'])

    return df


def load_acx_report(filepath, catalog=None):
    """
    Load a single ACX monthly royalty report (new format, post-April 2024).
    
    Uses a hybrid matching strategy:
    1. First attempts exact ID match against catalog 'other_id' field
    2. Falls back to fuzzy title matching for any unmatched rows
    
    Args:
        filepath: Path to the .xlsx file
        catalog: BookCatalog instance for identifier enrichment (optional)
    
    Returns:
        Normalized DataFrame with standardized column names and catalog info
    """
    filepath = Path(filepath)

    report_date = extract_date_from_acx_filename(filepath.name)
    print(f"[INFO] Loading ACX report for {report_date.strftime('%B %Y')} from {filepath.name}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if 'sales detail' in name.lower() and 'net sales' in name.lower():
            sheet_name = name
            break

    if sheet_name is None:
        raise ValueError(f"Could not find 'sales detail (net sales)' sheet in {filepath.name}")

    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' appears to be empty")

    headers = list(rows[0])
    data_rows = [list(r) for r in rows[1:] if any(cell is not None for cell in r)]

    wb.close()

    df = pd.DataFrame(data_rows, columns=headers)
    print(f"[INFO] Read {len(df)} rows from sheet '{sheet_name}'")

    df['sale_date'] = report_date

    df = df.rename(columns=ACX_NEW_FORMAT)

    for col in DROP_COLUMNS_COMMON:
        if col in df.columns:
            df = df.drop(columns=[col])
            print(f"[INFO] Dropped PII column: '{col}'")

    numeric_cols = ['quantity', 'price', 'royalty_amount', 'royalty_rate']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    if catalog is not None and 'book_identifier' in df.columns:
        print("[INFO] Stage 1: Attempting exact ID match against catalog...")
        df_enriched = catalog.enrich(df, 'book_identifier')

        matched_count = df_enriched['series'].notna().sum()
        unmatched_count = len(df_enriched) - matched_count

        if unmatched_count > 0:
            print(f"[INFO] Stage 1 matched {matched_count}/{len(df_enriched)} rows")
            print(f"[INFO] Stage 2: Falling back to title matching for {unmatched_count} unmatched rows...")

            unmatched_mask = df_enriched['series'].isna()
            unmatched_rows = df_enriched[unmatched_mask].copy()

            for col in ['series', 'canonical_work_slug', 'edition_format']:
                if col in unmatched_rows.columns:
                    unmatched_rows = unmatched_rows.drop(columns=[col])

            title_matched = _match_titles_to_catalog(unmatched_rows, catalog)

            matched_rows = df_enriched[~unmatched_mask]
            df_enriched = pd.concat([matched_rows, title_matched], ignore_index=True)

            final_matched = df_enriched['series'].notna().sum()
            print(f"[INFO] Hybrid enrichment complete: {final_matched}/{len(df_enriched)} rows matched")
        else:
            print(f"[INFO] Exact ID match succeeded for all {matched_count} rows")
    else:
        df_enriched = df.copy()
        df_enriched['series'] = None
        df_enriched['canonical_work_slug'] = None
        df_enriched['edition_format'] = None

    df = df_enriched

    df['source_platform'] = 'acx'

    print(f"[INFO] ACX load complete: {len(df)} normalized records")
    return df


def load_all_acx_reports(folder_path, catalog=None):
    """
    Batch process all ACX Excel files in a given folder.
    
    Args:
        folder_path: Path to folder containing ACX .xlsx files
        catalog: BookCatalog instance for enrichment
    
    Returns:
        Single combined DataFrame with all months concatenated
    """
    folder_path = Path(folder_path)

    acx_files = list(folder_path.glob('*.xlsx'))

    if not acx_files:
        print(f"[WARN] No .xlsx files found in {folder_path}")
        return pd.DataFrame()

    print(f"\n{'=' * 60}")
    print(f"BATCH PROCESSING: Found {len(acx_files)} ACX files")
    print(f"{'=' * 60}")

    all_dataframes = []

    for file_path in sorted(acx_files):
        try:
            df = load_acx_report(file_path, catalog=catalog)
            all_dataframes.append(df)
            print(f"  OK: {file_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"  FAILED: {file_path.name} - {e}")

    if not all_dataframes:
        print("[ERROR] No files processed successfully")
        return pd.DataFrame()

    combined = pd.concat(all_dataframes, ignore_index=True)

    print(f"\n{'=' * 60}")
    print(f"BATCH COMPLETE: {len(combined)} total rows from {len(all_dataframes)} files")
    print(f"Date range: {combined['sale_date'].min().strftime('%B %Y')} to {combined['sale_date'].max().strftime('%B %Y')}")
    print(f"{'=' * 60}")

    return combined


def load_acx_legacy_report(filepath, catalog=None):
    """
    Load a legacy ACX monthly royalty report (pre-April 2024, .xls format).
    
    Legacy format uses a cross-tab matrix (books as rows, regions as columns).
    This function unpivots the matrix into the same long format as the new loader.
    
    Args:
        filepath: Path to the .xls file
        catalog: BookCatalog instance for identifier enrichment (optional)
    
    Returns:
        Normalized DataFrame with standardized column names and catalog info
    """
    filepath = Path(filepath)

    report_date = extract_date_from_acx_filename(filepath.name)
    print(f"[INFO] Loading legacy ACX report for {report_date.strftime('%B %Y')} from {filepath.name}")

    df_raw = pd.read_excel(filepath, engine='xlrd', sheet_name='Summary', header=4)

    df_subtotals = df_raw[df_raw['Customer Type'] == 'Subtotal'].copy()

    df_subtotals['book_identifier'] = df_subtotals['Royalty Earner'].str.replace(
        'ACX-', '', regex=False
    )

    region_cols = ['US', 'UK', 'DE', 'FR', 'AU', 'CA', 'JP', 'IN', 'ES']
    region_cols = [c for c in region_cols if c in df_subtotals.columns]

    df_long = df_subtotals.melt(
        id_vars=['book_identifier'],
        value_vars=region_cols,
        var_name='region',
        value_name='royalty_amount'
    )

    df_long = df_long[
        df_long['royalty_amount'].notna() & (df_long['royalty_amount'] > 0)
    ].copy()

    df_long['sale_date'] = report_date
    df_long['source_platform'] = 'acx'
    df_long['quantity'] = None
    df_long['price'] = None
    df_long['royalty_rate'] = None

    if catalog is not None:
        print("[INFO] Stage 1: Attempting exact ID match against catalog...")
        df_enriched = catalog.enrich(df_long, 'book_identifier')

        matched_count = df_enriched['series'].notna().sum()
        unmatched_count = len(df_enriched) - matched_count

        if unmatched_count > 0:
            print(f"[INFO] {matched_count}/{len(df_enriched)} matched by ID")
            print(f"[INFO] {unmatched_count} unmatched (no book titles available for fuzzy fallback)")
        else:
            print(f"[INFO] Exact ID match succeeded for all {matched_count} rows")
    else:
        df_enriched = df_long.copy()
        df_enriched['series'] = None
        df_enriched['canonical_work_slug'] = None
        df_enriched['edition_format'] = None

    df = df_enriched

    print(f"[INFO] Legacy ACX load complete: {len(df)} normalized records")
    return df


def load_all_acx_legacy_reports(folder_path, catalog=None):
    """
    Batch process all legacy ACX .xls files in a given folder.
    
    Args:
        folder_path: Path to folder containing legacy ACX .xls files
        catalog: BookCatalog instance for enrichment
    
    Returns:
        Single combined DataFrame with all months concatenated
    """
    folder_path = Path(folder_path)

    legacy_files = list(folder_path.glob('*.xls'))

    if not legacy_files:
        print(f"[WARN] No .xls files found in {folder_path}")
        return pd.DataFrame()

    print(f"\n{'=' * 60}")
    print(f"BATCH PROCESSING (LEGACY): Found {len(legacy_files)} ACX files")
    print(f"{'=' * 60}")

    all_dataframes = []

    for file_path in sorted(legacy_files):
        try:
            df = load_acx_legacy_report(file_path, catalog=catalog)
            all_dataframes.append(df)
            print(f"  OK: {file_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"  FAILED: {file_path.name} - {e}")

    if not all_dataframes:
        print("[ERROR] No files processed successfully")
        return pd.DataFrame()

    combined = pd.concat(all_dataframes, ignore_index=True)

    print(f"\n{'=' * 60}")
    print(f"BATCH COMPLETE: {len(combined)} total rows from {len(all_dataframes)} files")
    print(f"Date range: {combined['sale_date'].min().strftime('%B %Y')} to {combined['sale_date'].max().strftime('%B %Y')}")
    print(f"{'=' * 60}")

    return combined