"""
Platform-specific data loaders for Author Business Analytics ETL pipeline.

Provides loader functions for extracting royalty data from Amazon publishing platforms,
normalizing column names to standardized schemas, enriching with book catalog
metadata, and returning DataFrames ready for ingestion into SQLite fact tables.

Supported Platforms & Formats:
- ACX (Audiobook Creation Exchange): Two loaders supporting both
  - New format (Apr 2024+): Multi-sheet .xlsx with "sales detail" sheets
  - Legacy format (Pre-April 2024): Cross-tab matrix in .xls (unpivot required)
- Amazon KDP: Combined Sales sheet from .xlsx exports

Enrichment Strategy:
1. Identifier-first matching (ACX ID, ASIN, ISBN)
2. Fallback to fuzzy title matching via _match_titles_to_catalog (shared utility)

Grain Declarations:
- ACX new format: One row per transaction-line item
- ACX legacy: One row per book-region aggregated monthly
- KDP: One row per transaction-line item (per book/format/sale)

Kimball Compliance:
Each platform maintains its own fact table due to incompatible grains,
measurements (sales vs pages vs subscription aggregations), and
platform-specific dimensions (region, distributor, fee chains).
"""

import calendar
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.schemas import (
    ACX_NEW_FORMAT,
    DROP_COLUMNS_COMMON,
    KDP_COMBINED_FORMAT,
    KDP_MARKETPLACE_MAP,
)
from src.matching import _match_titles_to_catalog


# ===================================================================
# UTILITY FUNCTIONS
# ===================================================================

def extract_date_from_acx_filename(filename: str) -> datetime:
    """
    Parse month and year from ACX report filename.

    Handles two naming conventions used by ACX reports sent over different periods:
    - New format (post-April 2024): Royalty_..._MONTHLY_Apr_2026.xlsx (abbreviated)
    - Old format (pre-April 2024): S_D_Simper_ACX_MONTHLY_April_2021.xls (full name)

    Args:
        filename: Filename string from ACX email attachment or folder listing

    Returns:
        datetime object representing the 1st day of the reported month/year

    Raises:
        ValueError: If the filename does not match known date patterns
    """
    # Build month lookup dictionary mapping lowercase names/abbreviations → numbers
    months = {}
    for i, name in enumerate(calendar.month_name[1:], 1):
        months[name.lower()] = i
    for i, abbr in enumerate(calendar.month_abbr[1:], 1):
        months[abbr.lower()] = i

    # Regex captures the date portion regardless of full vs abbreviated month
    pattern = r'_([A-Za-z]{3,9})_(\d{4})\.xlsx?$'
    match = re.search(pattern, filename, re.IGNORECASE)

    if match:
        month_str, year = match.groups()
        month_num = months.get(month_str.lower())
        if month_num:
            return datetime(int(year), month_num, 1)

    raise ValueError(f"Cannot parse date from filename: {filename}")


# ===================================================================
# ACX LOADERS (NEW FORMAT — POST APRIL 2024)
# ===================================================================

def load_acx_report(filepath: str | Path, catalog=None) -> pd.DataFrame:
    """
    Load a single ACX monthly royalty report (new format, post-April 2024).

    Parses multi-sheet .xlsx workbooks to locate the "Sales Detail (Net Sales)"
    sheet, extracts rows as-is, renames columns per ACX_NEW_FORMAT schema, drops
    PII columns defined in DROP_COLUMNS_COMMON, converts numeric fields, and
    performs hybrid catalog enrichment (identifier-first → title fallback).

    Grain: One row per transaction-line item sold through ACX distribution channels.

    Args:
        filepath: Path to the .xlsx file from ACX email/folder
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Normalized DataFrame with standardized column names and catalog info

    Raises:
        ValueError: If the specified sheet cannot be found or appears empty
    """
    filepath = Path(filepath)
    report_date = extract_date_from_acx_filename(filepath.name)
    print(f"[INFO] Loading ACX report for {report_date.strftime('%B %Y')} from {filepath.name}")

    # Open workbook in read-only mode with data_only=True to get calculated values
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Find sheet containing transactional data (sheet name varies slightly)
    sheet_name = None
    for name in wb.sheetnames:
        if 'sales detail' in name.lower() and 'net sales' in name.lower():
            sheet_name = name
            break

    if sheet_name is None:
        wb.close()
        raise ValueError(f"Could not find 'sales detail (net sales)' sheet in {filepath.name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' appears to be empty")

    headers = list(rows[0])
    data_rows = [list(r) for r in rows[1:] if any(cell is not None for cell in r)]
    wb.close()

    df = pd.DataFrame(data_rows, columns=headers)
    print(f"[INFO] Read {len(df)} rows from sheet '{sheet_name}'")

    df['sale_date'] = report_date
    df = df.rename(columns=ACX_NEW_FORMAT)

    for col in DROP_COLUMNS_COMMON:
        if col in df.columns:
            df = df.drop(columns=[col])
            print(f"[INFO] Dropped PII column: '{col}'")

    numeric_cols = ['quantity', 'price', 'royalty_amount', 'royalty_rate']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Hybrid enrichment: identifier-first, then fallback to fuzzy title matching
    if catalog is not None and 'book_identifier' in df.columns:
        print("[INFO] Stage 1: Attempting exact ID match against catalog...")
        df_enriched = catalog.enrich(df, 'book_identifier')

        matched_count = df_enriched['series'].notna().sum()
        unmatched_count = len(df_enriched) - matched_count

        if unmatched_count > 0:
            print(f"[INFO] Stage 1 matched {matched_count}/{len(df_enriched)} rows")
            print(f"[INFO] Stage 2: Falling back to title matching for {unmatched_count} unmatched rows...")

            unmatched_mask = df_enriched['series'].isna()
            unmatched_rows = df_enriched[unmatched_mask].copy()

            for col in ['series', 'canonical_work_slug', 'edition_format']:
                if col in unmatched_rows.columns:
                    unmatched_rows = unmatched_rows.drop(columns=[col])

            title_matched = _match_titles_to_catalog(unmatched_rows, catalog)

            matched_rows = df_enriched[~unmatched_mask]
            df_enriched = pd.concat([matched_rows, title_matched], ignore_index=True)

            final_matched = df_enriched['series'].notna().sum()
            print(f"[INFO] Hybrid enrichment complete: {final_matched}/{len(df_enriched)} rows matched")
        else:
            print(f"[INFO] Exact ID match succeeded for all {matched_count} rows")
    else:
        df_enriched = df.copy()
        df_enriched['series'] = None
        df_enriched['canonical_work_slug'] = None
        df_enriched['edition_format'] = None

    df = df_enriched
    df['source_platform'] = 'acx'

    print(f"[INFO] ACX load complete: {len(df)} normalized records")
    return df


def load_all_acx_reports(folder_path: str | Path, catalog=None) -> pd.DataFrame:
    """
    Batch process all ACX Excel files in a given folder.

    Iterates through all .xlsx files alphabetically, attempting to load each
    individually. Collects successful results and concatenates them after all
    files have been processed. Failed loads are logged but do not abort
    the entire batch run.

    Grain: Aggregated result set spanning multiple monthly export files.

    Args:
        folder_path: Path to folder containing ACX .xlsx files
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Single combined DataFrame with all months concatenated
    """
    folder_path = Path(folder_path)
    acx_files = list(folder_path.glob('*.xlsx'))

    if not acx_files:
        print(f"[WARN] No .xlsx files found in {folder_path}")
        return pd.DataFrame()

    print(f"\n{'=' * 60}")
    print(f"BATCH PROCESSING: Found {len(acx_files)} ACX files")
    print(f"{'=' * 60}")

    all_dataframes = []

    for file_path in sorted(acx_files):
        try:
            df = load_acx_report(file_path, catalog=catalog)
            all_dataframes.append(df)
            print(f"  OK: {file_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"  FAILED: {file_path.name} - {e}")

    if not all_dataframes:
        print("[ERROR] No files processed successfully")
        return pd.DataFrame()

    combined = pd.concat(all_dataframes, ignore_index=True)

    print(f"\n{'=' * 60}")
    print(f"BATCH COMPLETE: {len(combined)} total rows from {len(all_dataframes)} files")
    print(f"Date range: {combined['sale_date'].min().strftime('%B %Y')} to {combined['sale_date'].max().strftime('%B %Y')}")
    print(f"{'=' * 60}")

    return combined


# ===================================================================
# ACX LOADERS (LEGACY FORMAT — PRE APRIL 2024)
# ===================================================================

def load_acx_legacy_report(filepath: str | Path, catalog=None) -> pd.DataFrame:
    """
    Load a legacy ACX monthly royalty report (pre-April 2024, .xls format).

    Legacy format uses a cross-tab matrix structure where books appear as rows
    and regions as columns (US, UK, DE, FR, AU, CA, JP, IN, ES). This function
    identifies subtotal rows, extracts only those rows, melts/wide-to-long pivot
    to create one row per book-region combination, applies ID-based catalog
    enrichment (titles unavailable in legacy format), and adds placeholder
    N/A values for quantity, price, and royalty rate columns that don't exist.

    Grain: One row per book-region-month aggregation (not per-transaction).
    Kimball note: Requires separate fact table due to aggregate grain and lack
                  of transaction details necessary for joining with sale_fact.

    Args:
        filepath: Path to the .xls file from legacy ACX emails
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Normalized long-format DataFrame with standardized columns and catalog info

    Raises:
        ValueError: If Summary sheet cannot be parsed or contains no data
    """
    filepath = Path(filepath)
    report_date = extract_date_from_acx_filename(filepath.name)
    print(f"[INFO] Loading legacy ACX report for {report_date.strftime('%B %Y')} from {filepath.name}")

    df_raw = pd.read_excel(filepath, engine='xlrd', sheet_name='Summary', header=4)
    df_subtotals = df_raw[df_raw['Customer Type'] == 'Subtotal'].copy()

    # Extract ACX ID from "Royalty Earner" field (format: "ACX-XXXXXX")
    df_subtotals['book_identifier'] = df_subtotals['Royalty Earner'].str.replace(
        'ACX-', '', regex=False
    )

    region_cols = ['US', 'UK', 'DE', 'FR', 'AU', 'CA', 'JP', 'IN', 'ES']
    region_cols = [c for c in region_cols if c in df_subtotals.columns]

    # Unpivot cross-tab matrix to long format (one row per book-region)
    df_long = df_subtotals.melt(
        id_vars=['book_identifier'],
        value_vars=region_cols,
        var_name='region',
        value_name='royalty_amount'
    )

    # Filter out zero/null amounts (noise in cross-tabs)
    df_long = df_long[
        df_long['royalty_amount'].notna() & (df_long['royalty_amount'] > 0)
    ].copy()

    df_long['sale_date'] = report_date
    df_long['source_platform'] = 'acx'
    df_long['quantity'] = None
    df_long['price'] = None
    df_long['royalty_rate'] = None

    if catalog is not None:
        print("[INFO] Stage 1: Attempting exact ID match against catalog...")
        df_enriched = catalog.enrich(df_long, 'book_identifier')

        matched_count = df_enriched['series'].notna().sum()
        unmatched_count = len(df_enriched) - matched_count

        if unmatched_count > 0:
            print(f"[INFO] {matched_count}/{len(df_enriched)} matched by ID")
            print(f"[INFO] {unmatched_count} unmatched (legacy format provides no titles for fuzzy fallback)")
        else:
            print(f"[INFO] Exact ID match succeeded for all {matched_count} rows")
    else:
        df_enriched = df_long.copy()
        df_enriched['series'] = None
        df_enriched['canonical_work_slug'] = None
        df_enriched['edition_format'] = None

    df = df_enriched
    print(f"[INFO] Legacy ACX load complete: {len(df)} normalized records")
    return df


def load_all_acx_legacy_reports(folder_path: str | Path, catalog=None) -> pd.DataFrame:
    """
    Batch process all legacy ACX .xls files in a given folder.

    Follows same pattern as load_all_acx_reports but for pre-April 2024 .xls files.
    Graceful error handling ensures one corrupt file doesn't abort the batch.

    Args:
        folder_path: Path to folder containing legacy ACX .xls files
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Single combined DataFrame with all months concatenated
    """
    folder_path = Path(folder_path)
    legacy_files = list(folder_path.glob('*.xls'))

    if not legacy_files:
        print(f"[WARN] No .xls files found in {folder_path}")
        return pd.DataFrame()

    print(f"\n{'=' * 60}")
    print(f"BATCH PROCESSING (LEGACY): Found {len(legacy_files)} ACX files")
    print(f"{'=' * 60}")

    all_dataframes = []

    for file_path in sorted(legacy_files):
        try:
            df = load_acx_legacy_report(file_path, catalog=catalog)
            all_dataframes.append(df)
            print(f"  OK: {file_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"  FAILED: {file_path.name} - {e}")

    if not all_dataframes:
        print("[ERROR] No files processed successfully")
        return pd.DataFrame()

    combined = pd.concat(all_dataframes, ignore_index=True)

    print(f"\n{'=' * 60}")
    print(f"BATCH COMPLETE: {len(combined)} total rows from {len(all_dataframes)} files")
    print(f"Date range: {combined['sale_date'].min().strftime('%B %Y')} to {combined['sale_date'].max().strftime('%B %Y')}")
    print(f"{'=' * 60}")

    return combined


# ===================================================================
# AMAZON KDP LOADERS
# ===================================================================

def load_kdp_report(filepath: str | Path, catalog=None) -> pd.DataFrame:
    """
    Load a single Amazon KDP royalty report.

    Reads the 'Combined Sales' sheet from KDP .xlsx exports, which contains
    all book formats (eBook, Paperback, Hardcover) in a flat tabular structure.
    Applies schema normalization via KDP_COMBINED_FORMAT mappings, drops PII
    columns, transforms marketplace domains to region codes, parses royalty
    type percentages to decimals, and performs hybrid catalog enrichment.

    Grain: One row per transaction-line item (per book/format/sale).

    Args:
        filepath: Path to the .xlsx file from KDP Reports portal
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Normalized DataFrame with standardized column names and catalog info

    Raises:
        ValueError: If 'Combined Sales' sheet cannot be found or parsed
    """
    filepath = Path(filepath)
    print(f"[INFO] Loading KDP report from {filepath.name}")

    df = pd.read_excel(filepath, sheet_name='Combined Sales', engine='openpyxl')
    print(f"[INFO] Read {len(df)} rows from 'Combined Sales'")

    # Step 1: Apply schema normalization
    df = df.rename(columns=KDP_COMBINED_FORMAT)

    # Step 2: Drop PII columns
    for col in DROP_COLUMNS_COMMON:
        if col in df.columns:
            df = df.drop(columns=[col])
            print(f"[INFO] Dropped PII column: '{col}'")

    # Step 3: Transform marketplace domains to region codes
    df['region'] = df['marketplace_raw'].map(KDP_MARKETPLACE_MAP)
    df = df.drop(columns=['marketplace_raw'])

    # Step 4: Parse royalty type percentage (e.g., "35%" → 0.35)
    df['royalty_rate'] = df['royalty_type_raw'].astype(str).str.replace('%', '').astype(float) / 100
    df = df.drop(columns=['royalty_type_raw'])

    # Step 5: Parse sale_date from "YYYY-MM" string to datetime
    df['sale_date'] = pd.to_datetime(df['sale_date'], format='%Y-%m')

    # Step 6: Type conversion for numeric columns
    numeric_cols = ['quantity', 'price', 'royalty_amount']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Step 7: Enrich with catalog using hybrid strategy
    if catalog is not None and 'book_identifier' in df.columns:
        print("[INFO] Stage 1: Attempting exact ID match against catalog...")
        df_enriched = catalog.enrich(df, 'book_identifier')

        matched_count = df_enriched['series'].notna().sum()
        unmatched_count = len(df_enriched) - matched_count

        if unmatched_count > 0:
            print(f"[INFO] Stage 1 matched {matched_count}/{len(df_enriched)} rows")
            print(f"[INFO] Stage 2: Falling back to title matching for {unmatched_count} unmatched rows...")

            unmatched_mask = df_enriched['series'].isna()
            unmatched_rows = df_enriched[unmatched_mask].copy()

            for col in ['series', 'canonical_work_slug', 'edition_format']:
                if col in unmatched_rows.columns:
                    unmatched_rows = unmatched_rows.drop(columns=[col])

            title_matched = _match_titles_to_catalog(unmatched_rows, catalog)

            matched_rows = df_enriched[~unmatched_mask]
            df_enriched = pd.concat([matched_rows, title_matched], ignore_index=True)

            final_matched = df_enriched['series'].notna().sum()
            print(f"[INFO] Hybrid enrichment complete: {final_matched}/{len(df_enriched)} rows matched")
        else:
            print(f"[INFO] Exact ID match succeeded for all {matched_count} rows")
    else:
        df_enriched = df.copy()
        df_enriched['series'] = None
        df_enriched['canonical_work_slug'] = None
        df_enriched['edition_format'] = None

    df = df_enriched
    df['source_platform'] = 'amazon_kdp'

    print(f"[INFO] KDP load complete: {len(df)} normalized records")
    return df


def load_all_kdp_reports(folder_path: str | Path, catalog=None) -> pd.DataFrame:
    """
    Batch process all Amazon KDP Excel files in a given folder.

    Iterates through all .xlsx files, logs progress and errors per-file, and
    concatenates successful loads into a single master DataFrame.

    Args:
        folder_path: Path to folder containing KDP .xlsx files
        catalog: Optional BookCatalog instance for enrichment

    Returns:
        Single combined DataFrame with all files concatenated
    """
    folder_path = Path(folder_path)
    kdp_files = list(folder_path.glob('*.xlsx'))

    if not kdp_files:
        print(f"[WARN] No .xlsx files found in {folder_path}")
        return pd.DataFrame()

    print(f"\n{'=' * 60}")
    print(f"BATCH PROCESSING (KDP): Found {len(kdp_files)} files")
    print(f"{'=' * 60}")

    all_dataframes = []

    for file_path in sorted(kdp_files):
        try:
            df = load_kdp_report(file_path, catalog=catalog)
            all_dataframes.append(df)
            print(f"  OK: {file_path.name} ({len(df)} rows)")
        except Exception as e:
            print(f"  FAILED: {file_path.name} - {e}")

    if not all_dataframes:
        print("[ERROR] No files processed successfully")
        return pd.DataFrame()

    combined = pd.concat(all_dataframes, ignore_index=True)

    print(f"\n{'=' * 60}")
    print(f"BATCH COMPLETE: {len(combined)} total rows from {len(all_dataframes)} files")
    print(f"Date range: {combined['sale_date'].min().strftime('%B %Y')} to {combined['sale_date'].max().strftime('%B %Y')}")
    print(f"{'=' * 60}")

    return combined